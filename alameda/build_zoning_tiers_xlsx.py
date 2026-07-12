#!/usr/bin/env python3
"""Build zoning shelter/TH tier crosswalk Excel from municipal-code research."""

from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/bob-brown/Downloads/tvs-ai/alameda/zoning_shelter_tiers.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "zoning_tiers"

headers = [
    "city",
    "base_zone",
    "tier",
    "allows_shelter_by_right",
    "allows_transitional_housing",
    "notes",
]
ws.append(headers)

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF")
tier_fills = {
    "A": PatternFill("solid", fgColor="C6EFCE"),
    "B": PatternFill("solid", fgColor="FFEB9C"),
    "C": PatternFill("solid", fgColor="FFC7CE"),
}
thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
for col, _h in enumerate(headers, 1):
    cell = ws.cell(1, col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(wrap_text=True, vertical="center")

rows: list[tuple] = []


def add(city, zone, tier, shelter, th, notes):
    rows.append((city, zone, tier, shelter, th, notes))


# =============================================================================
# OAKLAND — Planning Code Title 17 (17.103.010/015 + zone chapters)
# =============================================================================

for z in ["CIX-1", "CIX-2", "IG", "IO"]:
    note = (
        "SB2 shelter zone: by-right on church properties and on 3rd St / E 12th / "
        "Coliseum Way corridor segments (OPC 17.103.015 / 17.73 L1); other residential "
        "generally prohibited"
    )
    th = "no"
    if z == "CIX-2":
        note += (
            "; motel→TH/SH conversion may be CUP if >1,500 ft from Hegenberger"
        )
    add("Oakland", z, "A", "yes", th, note)

for z in ["HBX-1", "HBX-2", "HBX-3"]:
    add(
        "Oakland",
        z,
        "B",
        "no",
        "yes",
        "TH/SH permitted as residential (P); shelter by-right on churches + E 12th corridor "
        "segment, CUP elsewhere (OPC 17.65 L2)",
    )

for z in ["CN-1", "CN-2", "CN-3", "CN-4"]:
    add(
        "Oakland",
        z,
        "B",
        "no",
        "yes",
        "Shelter by-right only on churches + MLK/San Pablo/MacArthur corridor segments; "
        "CUP elsewhere (OPC 17.33 L5); TH/SH permitted with ground-floor limits",
    )

for z, th, extra in [
    ("CC-1", "yes", "TH/SH permitted (P) with limits"),
    ("CC-2", "yes", "TH/SH permitted (P) with limits"),
    ("CC-3", "no", "TH/SH require CUP (C)"),
]:
    add(
        "Oakland",
        z,
        "B",
        "no",
        th,
        "Shelter by-right only on churches + San Pablo/Webster corridor segments; "
        f"CUP elsewhere (OPC 17.35 L4); {extra}",
    )

add(
    "Oakland",
    "CR-1",
    "B",
    "no",
    "yes",
    "Shelter by-right on churches / limited corridor path; TH/SH P in applicable CR "
    "subzone (OPC 17.37); CUP or corridor path otherwise",
)

for z in ["CBD-P", "CBD-C", "CBD-R", "CBD-X"]:
    add(
        "Oakland",
        z,
        "B",
        "no",
        "yes",
        "TH/SH treated as residential where permanent residential allowed; shelter "
        "by-right on churches + designated corridors, CUP elsewhere (OPC CBD / 17.103)",
    )

for z in ["RH-1", "RH-2", "RH-3", "RH-4"]:
    add(
        "Oakland",
        z,
        "B",
        "no",
        "yes",
        "TH/SH permitted as residential (P); emergency shelter by-right only on church "
        "properties, prohibited elsewhere (OPC 17.13 L1)",
    )

for z in ["RD-1", "RD-2"]:
    add(
        "Oakland",
        z,
        "B",
        "no",
        "yes",
        "TH/SH permitted as residential (P); emergency shelter by-right only on church "
        "properties, prohibited elsewhere (OPC 17.15 L1)",
    )

for z in ["RM-1", "RM-2", "RM-3", "RM-4"]:
    add(
        "Oakland",
        z,
        "B",
        "no",
        "yes",
        "TH/SH permitted as residential (P); shelter by-right only on churches + MLK "
        "corridor segment, prohibited elsewhere (OPC 17.17 L2)",
    )

for z in ["RU-1", "RU-2", "RU-3", "RU-4", "RU-5"]:
    add(
        "Oakland",
        z,
        "B",
        "no",
        "yes",
        "TH/SH permitted as residential (P); shelter by-right on churches + "
        "MLK/San Pablo/MacArthur corridors, CUP elsewhere (OPC 17.19 L2)",
    )

add(
    "Oakland",
    "S-15",
    "B",
    "no",
    "yes",
    "TH/SH permitted (P); emergency shelter P with 17.103 limits "
    "(by-right churches/corridors; CUP path elsewhere)",
)

for z, note in [
    ("S-1", "Medical center / special; shelter/TH not generally by-right residential"),
    ("S-2", "Civic / special district; village use typically needs discretionary review or rezone"),
    ("S-3", "Research / special; village use typically needs discretionary review or rezone"),
]:
    add("Oakland", z, "C", "no", "no", note)

for z, tier, sh, th, note in [
    (
        "C-40",
        "B",
        "no",
        "yes",
        "Legacy commercial; CUP/corridor shelter path; TH where residential allowed",
    ),
    (
        "C-45",
        "B",
        "no",
        "yes",
        "Legacy commercial; CUP/corridor shelter path; TH where residential allowed",
    ),
    (
        "M-20",
        "C",
        "no",
        "no",
        "Legacy industrial; not an SB2 by-right shelter district; residential generally prohibited",
    ),
    (
        "M-30",
        "C",
        "no",
        "no",
        "Legacy industrial; not an SB2 by-right shelter district; residential generally prohibited",
    ),
    (
        "M-40",
        "C",
        "no",
        "no",
        "Legacy industrial; not an SB2 by-right shelter district; residential generally prohibited",
    ),
    (
        "R-80",
        "B",
        "no",
        "yes",
        "Legacy high-density residential; TH/SH as residential; shelter not zone-wide by-right",
    ),
    (
        "D-KP-1",
        "B",
        "no",
        "yes",
        "Kaiser Permanente district; TH/SH P with limits; shelter P(L) per OPC 17.101D",
    ),
    (
        "D-KP-2",
        "B",
        "no",
        "yes",
        "Kaiser Permanente district; TH/SH P with limits; shelter P(L) per OPC 17.101D",
    ),
    (
        "D-KP-3",
        "B",
        "no",
        "yes",
        "Kaiser Permanente district; TH/SH P with limits; shelter P(L) per OPC 17.101D",
    ),
    (
        "D-KP-4",
        "B",
        "no",
        "yes",
        "Kaiser Permanente D-KP-4; RU-3-like limits for residual residential parcels",
    ),
    (
        "K-DP-4",
        "B",
        "no",
        "yes",
        "GIS label variant of D-KP-4 / Kaiser Permanente district",
    ),
    (
        "D-OTN",
        "B",
        "no",
        "yes",
        "Oak-to-Ninth / Brooklyn Basin; TH/SH P in D-OTN; shelter P(L)",
    ),
    (
        "Wood Street",
        "B",
        "no",
        "yes",
        "Wood Street Development; residential-oriented specific plan — TH/SH as residential; "
        "shelter not zone-wide SB2",
    ),
]:
    add("Oakland", z, tier, sh, th, note)

for z in [
    "OS (NP)",
    "OS(NP)",
    "OS (LP)",
    "OS(LP)",
    "OS (CP)",
    "OS(CP)",
    "OS (AF)",
    "OS (AMP)",
    "OS (PMP)",
    "OS(PMP)",
    "OS (RCA)",
    "OS (RSP)",
    "OS (SU)",
]:
    add(
        "Oakland",
        z,
        "C",
        "no",
        "no",
        "Open space district; village use not permitted without rezone (excluded)",
    )

add(
    "Oakland",
    "RU-3/D-BR",
    "B",
    "no",
    "yes",
    "GIS basezone encodes Broadway overlay with RU-3; same tier as RU-3 "
    "(TH/SH P; shelter CUP/corridor — D-BR does not change shelter/TH flags)",
)

# Oakland S/C/D overlays do not change shelter/TH columns vs base → omitted

# =============================================================================
# BERKELEY — BMC Table 23.308-1; supportive housing 23.302.070
# =============================================================================

berk_commercial_zc = {
    "C-1": "General Commercial (C-1); shelter ≤25 beds ZC (BMC 23.308-1); TH/SH as residential",
    "C-N": "Neighborhood Commercial; shelter ≤25 beds ZC; TH/SH as residential",
    "C-N(H)": "C-N with Hillside; same shelter/TH rules as C-N (H encoded in GIS zoneclass)",
    "C-E": "Elmwood Commercial; shelter ≤25 beds ZC; TH/SH as residential",
    "C-NS": "North Shattuck Commercial; shelter ≤25 beds ZC; TH/SH as residential",
    "C-NS(H)": "C-NS with Hillside; same shelter/TH as C-NS",
    "C-SA": "South Area Commercial; shelter ≤25 beds ZC; TH/SH as residential",
    "C-T": "Telegraph Commercial; shelter ≤25 beds ZC; TH/SH as residential",
    "C-SO": "Solano Commercial; shelter ≤25 beds ZC; TH/SH as residential",
    "C-W": "West Berkeley Commercial; shelter ≤25 beds ZC; TH/SH as residential",
    "C-AC": "Adeline Corridor Commercial; shelter ≤25 beds ZC; TH/SH as residential",
}
for z, note in berk_commercial_zc.items():
    add("Berkeley", z, "A", "yes", "yes", note + "; >bed cap needs UP(PH)")

for z, label in [
    ("C-DMU Core", "C-DMU Core"),
    ("C-DMU Buff", "C-DMU Buffer"),
    ("C-DMU Corr", "C-DMU Corridor"),
    ("C-DMU Oute", "C-DMU Outer"),
]:
    add(
        "Berkeley",
        z,
        "A",
        "yes",
        "yes",
        f"{label}: shelter ≤60 beds ZC (BMC 23.308-1); TH/SH as residential; >60 beds UP(PH)",
    )

for z in ["R-4", "R-4H", "R-5", "R-5H", "R-S", "R-SH", "R-SMU"]:
    add(
        "Berkeley",
        z,
        "A",
        "yes",
        "yes",
        "Shelter ≤15 beds ZC (BMC 23.308-1); TH/SH as residential / supportive housing "
        "by-right where multifamily allowed; >15 beds UP(PH)",
    )

for z in [
    "R-1",
    "R-1A",
    "R-1H",
    "R-2",
    "R-2A",
    "R-2AH",
    "R-2H",
    "R-3",
    "R-3H",
    "ES-R",
]:
    add(
        "Berkeley",
        z,
        "B",
        "no",
        "yes",
        "Year-round emergency shelter not permitted (BMC 23.308-1); TH/SH allowed as "
        "residential use / supportive housing pathway; seasonal shelter only with limits "
        "incidental to community/institutional use",
    )

add(
    "Berkeley",
    "MUR",
    "B",
    "no",
    "yes",
    "MU-R Mixed Use-Residential (GIS: MUR): emergency shelter not permitted "
    "(BMC 23.308-1); TH/SH as residential in MU-R",
)
add(
    "Berkeley",
    "MULI",
    "C",
    "no",
    "no",
    "MU-LI Mixed Use-Light Industrial (GIS: MULI): emergency shelter not permitted; "
    "residential/TH pathway not available without rezone (BMC 23.308-1)",
)
add(
    "Berkeley",
    "M",
    "C",
    "no",
    "no",
    "Manufacturing: emergency shelter not permitted; residential/TH not permitted "
    "(BMC 23.308-1)",
)
add(
    "Berkeley",
    "MM",
    "C",
    "no",
    "no",
    "Mixed Manufacturing: emergency shelter not permitted; residential/TH not permitted "
    "(BMC 23.308-1)",
)
add(
    "Berkeley",
    "U",
    "C",
    "no",
    "no",
    "Unclassified; village use not permitted without zoning determination/rezone",
)
add(
    "Berkeley",
    "X",
    "C",
    "no",
    "no",
    "Special/excluded mapping code; not a developable shelter/TH district",
)
add(
    "Berkeley",
    "SP",
    "B",
    "no",
    "yes",
    "Specific Plan area — check applicable SP use table; often allows residential/TH "
    "with plan standards",
)

# =============================================================================
# SAN LEANDRO — Zoning Code CC/IL/IG by-right; 4.04.316 TH/SH; S overlay → CUP
# =============================================================================

for z, sh, th, note in [
    (
        "CC",
        "yes",
        "yes",
        "SB2/commercial: emergency shelters ≤25 beds by-right; >25 beds CUP; TH/SH as "
        "residential (CC uses / 4.04.316)",
    ),
    (
        "IL",
        "yes",
        "no",
        "SB2 industrial: emergency shelters ≤45 beds by-right; >45 beds CUP; general "
        "residential/TH not permitted in IL",
    ),
    (
        "IG",
        "yes",
        "no",
        "SB2 industrial: emergency shelters ≤45 beds by-right; >45 beds CUP; general "
        "residential/TH not permitted in IG",
    ),
]:
    add("San Leandro", z, "A", sh, th, note)

add(
    "San Leandro",
    "IT",
    "B",
    "no",
    "no",
    "Emergency shelters conditionally permitted within 1/2 mile of BART; TH/SH not a "
    "standard residential allowance in IT",
)

for z, note in [
    (
        "CN",
        "Commercial Neighborhood: TH/SH as residential where residential allowed; "
        "emergency shelter not a by-right SB2 use (see CC/IL/IG)",
    ),
    (
        "CR",
        "Commercial Recreation: TH/SH where residential allowed; shelter not by-right SB2 district",
    ),
    (
        "CS",
        "Commercial Services: TH/SH where residential allowed; shelter not by-right SB2 district",
    ),
    (
        "P",
        "Professional Office: TH/SH where residential allowed; shelter not by-right SB2 district",
    ),
    (
        "DA-1",
        "Downtown Area retail mixed-use: TH/SH as residential/mixed-use; shelter not listed "
        "by-right like CC/IL/IG",
    ),
    (
        "DA-2",
        "Multi-use infill: TH/SH as residential/mixed-use; shelter not by-right SB2 district",
    ),
    (
        "DA-3",
        "TOD transition mixed-use: TH/SH as residential/mixed-use; shelter not by-right SB2 district",
    ),
    (
        "DA-4",
        "TOD residential mixed-use: TH/SH as residential/mixed-use; shelter not by-right SB2 district",
    ),
    (
        "DA-6",
        "Office mixed-use: TH/SH as residential/mixed-use; shelter not by-right SB2 district",
    ),
    (
        "NA-1",
        "North Area-1: TH/SH as residential where allowed; shelter not by-right SB2 district",
    ),
    (
        "NA-2",
        "North Area-2: TH/SH as residential where allowed; shelter not by-right SB2 district",
    ),
    (
        "SA-1",
        "South Area-1: TH/SH as residential where allowed; shelter not by-right SB2 district",
    ),
    (
        "SA-2",
        "South Area-2: TH/SH as residential where allowed; shelter not by-right SB2 district",
    ),
    (
        "SA-3",
        "South Area-3: TH/SH as residential where allowed; shelter not by-right SB2 district",
    ),
    (
        "B-TOD",
        "Bay Fair TOD: TH/SH as residential/mixed-use; shelter not the citywide SB2 by-right district",
    ),
]:
    add("San Leandro", z, "B", "no", "yes", note)

for z, desc in [
    ("RS", "Residential Single-Family"),
    ("RS-40", "Residential Single-Family 40-ft setback"),
    ("RS-VP", "Residential Single-Family view preservation"),
    ("RD", "Residential Duplex"),
    ("RO", "Residential Outer"),
    ("RM-1800", "Residential Multi-Family (24 du/ac)"),
    ("RM-2000", "Residential Multi-Family (22 du/ac)"),
    ("RM-2500", "Residential Multi-Family (17.5 du/ac)"),
    ("RM-3000", "Residential Multi-Family (14.5 du/ac)"),
]:
    add(
        "San Leandro",
        z,
        "B",
        "no",
        "yes",
        f"{desc}: TH/SH considered residential use (4.04.316); supportive housing ≤50 units "
        "by-right where multifamily allowed (Gov. Code 65651); emergency shelter not by-right",
    )

add(
    "San Leandro",
    "PS",
    "B",
    "no",
    "yes",
    "Public/Semipublic: institutional/shelter-crisis or public-facility pathway possible; "
    "not an SB2 by-right shelter district; TH/SH if residential component allowed",
)
add(
    "San Leandro",
    "IP",
    "C",
    "no",
    "no",
    "Industrial Park: emergency shelter not a permitted by-right use; residential/TH not "
    "permitted without rezone",
)
add(
    "San Leandro",
    "OS",
    "C",
    "no",
    "no",
    "Open Space: village use not permitted without rezone (excluded)",
)

# S overlay changes by-right shelter → discretionary
for z in ["CC", "IL", "IG"]:
    th = "yes" if z == "CC" else "no"
    add(
        "San Leandro",
        f"{z}/S",
        "B",
        "no",
        th,
        f"Special Review (S) overlay: base {z} by-right shelter may require CUP "
        "(Zoning Code S overlay note); treat as discretionary",
    )

add(
    "San Leandro",
    "IG/AU, S",
    "B",
    "no",
    "no",
    "IG with Assembly Use + Special Review: S overlay may require CUP for otherwise "
    "permitted shelter",
)
add(
    "San Leandro",
    "IL/AU, S",
    "B",
    "no",
    "no",
    "IL with Assembly Use + Special Review: S overlay may require CUP for otherwise "
    "permitted shelter",
)
add(
    "San Leandro",
    "CC/PD, S",
    "B",
    "no",
    "yes",
    "CC with PD + Special Review: S overlay may require CUP for otherwise by-right shelter",
)

# =============================================================================
# HAYWARD — HMC Ch. 10; HE App. D; MB Code
# =============================================================================

for z in ["IL", "IG", "IP"]:
    add(
        "Hayward",
        z,
        "A",
        "yes",
        "no",
        "SB2 industrial: Homeless/Emergency Shelter permitted by-right on city-owned "
        "property (HMC 10-1.1600); TH/SH not standard residential in industrial",
    )

for z in ["MB-CN", "MB-NN"]:
    add(
        "Hayward",
        z,
        "A",
        "yes",
        "yes",
        "SB2 / Mission Blvd Code: Emergency Homeless Shelter permitted in same manner "
        "as multifamily residential (HE App. D; MB Code); TH/SH as residential",
    )

for z, name in [
    ("CN", "Neighborhood Commercial"),
    ("CG", "General Commercial"),
    ("CO", "Commercial Office"),
]:
    add(
        "Hayward",
        z,
        "A",
        "yes",
        "yes",
        f"{name}: Emergency Shelter permitted by-right on parcels abutting Mission Blvd "
        "or Foothill Blvd (HMC 10-1.2736(a)); TH/SH as residential where residential allowed",
    )

add(
    "Hayward",
    "CN-R",
    "B",
    "no",
    "yes",
    "Neighborhood Commercial-Residential: Low Barrier Navigation Center permitted; "
    "emergency shelter not listed as Mission/Foothill by-right use like CN; TH/SH as residential",
)

for z, note in [
    (
        "CL",
        "Limited Access Commercial: TH/SH where residential allowed; not an SB2 by-right "
        "shelter district",
    ),
    (
        "CB",
        "Central Business: TH/SH where residential allowed; shelter not citywide SB2 "
        "by-right district",
    ),
    (
        "CBB20",
        "CB with B20 combining: same use permissions as CB for shelter/TH columns",
    ),
    (
        "CC-C",
        "Central City Commercial: TH/SH where residential allowed; shelter not SB2 "
        "by-right district",
    ),
    (
        "CC-R",
        "Central City Residential: TH/SH as residential; shelter not SB2 by-right district",
    ),
    (
        "RO",
        "Residential Office: TH/SH as residential; shelter not SB2 by-right district",
    ),
    (
        "SMU",
        "Sustainable Mixed Use: TH/SH as residential; shelter not SB2 by-right district",
    ),
    (
        "DT-MS",
        "Downtown Main Street: TH/SH as residential in Downtown Code; shelter not "
        "MB-CN/NN SB2 district",
    ),
    (
        "UN",
        "Urban Neighborhood: TH/SH as residential; shelter not SB2 by-right district",
    ),
    (
        "UNL",
        "Urban Neighborhood Limited: TH/SH as residential; shelter not SB2 by-right district",
    ),
    (
        "UC",
        "Urban Center: TH/SH as residential; shelter not SB2 by-right district",
    ),
    (
        "NE",
        "Neighborhood Edge: TH/SH as residential; shelter not SB2 by-right district",
    ),
    (
        "NG",
        "Neighborhood General: TH/SH as residential; shelter not SB2 by-right district",
    ),
    (
        "MB-CC",
        "Mission Blvd Corridor Center: multifamily/TH pathway; emergency shelter not "
        "unqualified by-right like MB-CN/NN (HE D-14)",
    ),
]:
    add("Hayward", z, "B", "no", "yes", note)

add(
    "Hayward",
    "MB-CS",
    "C",
    "no",
    "no",
    "Mission Blvd Civic Space: not a residential or SB2 shelter district; rezone needed "
    "for village use",
)

for z, desc in [
    ("RL", "Low Density Residential"),
    ("RLB4", "RL with B4 lot combining"),
    ("RLB6", "RL with B6 lot combining"),
    ("RLB8", "RL with B8 lot combining"),
    ("RLB10", "RL with B10 lot combining"),
    ("RLB20", "RL with B20 lot combining"),
    ("RLB40", "RL with B40 lot combining"),
    ("RM", "Medium Density Residential"),
    ("RMB3.5", "RM with B3.5 combining"),
    ("RMB4", "RM with B4 combining"),
    ("RH", "High Density Residential"),
    ("RHB7", "RH with B7 combining"),
    ("RNP", "Residential Natural Preserve"),
    ("MH", "Mobile Home Park"),
]:
    add(
        "Hayward",
        z,
        "B",
        "no",
        "yes",
        f"{desc}: TH/SH permitted as residential use (HMC / HE App. D); emergency shelter "
        "not by-right outside industrial/MB-CN/NN/Mission-Foothill commercial",
    )

for z in ["A", "AB10A", "AB160A"]:
    add(
        "Hayward",
        z,
        "C",
        "no",
        "no",
        "Agriculture district: village/shelter/TH not permitted without rezone",
    )

for z, note in [
    ("AT-AC", "Air Terminal Aviation Commercial: not an SB2 shelter/TH district"),
    ("AT-C", "Air Terminal Commercial: not an SB2 shelter/TH district"),
    ("AT-IP", "Air Terminal Industrial Park: not an SB2 shelter/TH district"),
    ("AT-O", "Air Terminal (other): not an SB2 shelter/TH district"),
    ("AT-R", "Air Terminal Recreation: not an SB2 shelter/TH district"),
    (
        "FP",
        "Flood Plain: development heavily constrained; village use not permitted without "
        "special approvals/rezone",
    ),
    ("OS", "Open Space: excluded; village use not permitted without rezone"),
    (
        "County",
        "Alameda County jurisdiction island/unincorporated mapping — City of Hayward "
        "zoning does not apply",
    ),
]:
    add("Hayward", z, "C", "no", "no", note)

add(
    "Hayward",
    "AT-RM",
    "B",
    "no",
    "yes",
    "Air Terminal Med Density Residential: TH/SH as residential if housing allowed; "
    "shelter not SB2 by-right",
)
add(
    "Hayward",
    "PF",
    "B",
    "no",
    "no",
    "Public Facilities: public/shelter-crisis pathway may exist but not SB2 by-right "
    "private shelter district; TH only if residential component authorized",
)
add(
    "Hayward",
    "PD",
    "B",
    "no",
    "yes",
    "Planned Development: use depends on PD approvals; not a standard by-right shelter "
    "district — site-specific; TH/SH possible if PD allows residential",
)

# Overlay that changes TH process risk on MB-CN
add(
    "Hayward",
    "MB-CN/Commercial Overlay 2",
    "B",
    "yes",
    "yes",
    "MB-CN with Commercial Overlay 2: shelter still allowed as in MB-CN, but residential "
    "along primary street frontage may require CUP (HE App. D note 4) — added process risk "
    "for TH/SH",
)

city_order = {"Oakland": 0, "Berkeley": 1, "San Leandro": 2, "Hayward": 3}
rows.sort(key=lambda r: (city_order.get(r[0], 9), r[2], r[1]))

for r in rows:
    ws.append(list(r))
    if r[2] in tier_fills:
        ws.cell(ws.max_row, 3).fill = tier_fills[r[2]]
    for c in range(1, 7):
        ws.cell(ws.max_row, c).border = thin
        ws.cell(ws.max_row, c).alignment = Alignment(wrap_text=True, vertical="top")

widths = [14, 36, 6, 22, 26, 80]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30
ws.auto_filter.ref = f"A1:F{ws.max_row}"
ws.freeze_panes = "A2"

ws2 = wb.create_sheet("sources_methodology")
ws2["A1"] = "Zoning tier crosswalk — sources & methodology"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A3"] = """
Tier definitions (project):
- A: Emergency shelter permitted without CUP (city SB2 / ministerial shelter zone), or effectively by-right under objective standards.
- B: Use allowed with CUP / discretionary review, or viable via TH/SH-as-residential or shelter-crisis / interim pathway.
- C: Village use not permitted without a rezone (e.g., open space, heavy industrial without shelter allowance).

Column rules:
- allows_shelter_by_right: yes only if the base district (or city-designated SB2 geography tied to that district) allows emergency shelter without CUP for qualifying projects (bed caps / objective standards may still apply).
- allows_transitional_housing: yes if transitional/supportive housing is allowed as a residential use in that district (SB2 / Gov. Code 65583).
- Overlay rows: only when the overlay changes tier / shelter / TH values vs the base alone. Otherwise automation should strip the overlay and use the base row.

Oakland sources:
- Oakland Planning Code Title 17, esp. 17.103.010, 17.103.015 (by-right corridors + churches); zone chapters 17.13–17.19, 17.33, 17.35, 17.37, 17.65, 17.73, CBD/D-KP/D-OTN/OS.
- City page: Emergency Shelter & Transitional Housing Zoning (SB2 implementation).

Berkeley sources:
- BMC Table 23.308-1 Permit Requirements for Emergency Shelters (ZC bed caps by district).
- BMC 23.302.070 Supportive Housing by-right where multifamily/mixed-use permitted.
- GIS zoneclass used as base_zone (H hillside encoded in zoneclass where present).

San Leandro sources:
- Zoning Code CC / IL / IG use lists (emergency shelters by bed count); IT CUP near BART.
- §4.04.316 Supportive and Transitional Housing; §4.04.384 Emergency Shelters.
- S Overlay: permitted uses may require CUP — combo rows for CC/S, IL/S, IG/S and multi-overlay variants.

Hayward sources:
- HMC industrial districts (shelter P* on city-owned); CN/CG/CO Mission Blvd & Foothill Blvd abutting parcels.
- Mission Boulevard Code MB-CN / MB-NN (shelter as multifamily).
- Housing Element Appendix D Tables D-13 / D-14 (TH/SH as residential; shelter geography).
- HMC 10-1.2736 emergency shelter standards (45-bed cap, 300-ft separation).

Notes / caveats:
- Oakland and Hayward SB2 geographies are partly corridor- or ownership-limited inside a zone; zone is still tagged A when that is the city's SB2 vehicle (matching CIX-1 example pattern).
- Parcel-level corridor / city-ownership checks still required in automation for true by-right eligibility.
- PD / Specific Plan sites may override; verify adopted PD/SP conditions.
- Built from municipal codes / housing elements; does not use the prior incorrect Zoning.xlsx.
""".strip()
ws2["A3"].alignment = Alignment(wrap_text=True, vertical="top")
ws2.column_dimensions["A"].width = 120
ws2.row_dimensions[3].height = 420

ws3 = wb.create_sheet("summary")
ws3.append(["city", "tier", "count"])
ctr = Counter((r[0], r[2]) for r in rows)
for (city, tier), n in sorted(
    ctr.items(), key=lambda x: (city_order.get(x[0][0], 9), x[0][1])
):
    ws3.append([city, tier, n])
ws3.append([])
ws3.append(["total_rows", len(rows)])

wb.save(OUT)
print(f"Wrote {OUT} with {len(rows)} rows")
print("Counts by city/tier:")
for (city, tier), n in sorted(
    ctr.items(), key=lambda x: (city_order.get(x[0][0], 9), x[0][1])
):
    print(f"  {city:12} {tier}: {n}")
print("\nKey Oakland samples:")
for r in rows:
    if r[0] == "Oakland" and r[1] in (
        "CIX-1",
        "RM-1",
        "OS (NP)",
        "IG",
        "RU-1",
        "CN-3",
        "HBX-1",
    ):
        print(f"  {r}")
